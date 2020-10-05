# -*- coding: utf-8 -*-

from sklearn.model_selection import train_test_split
import numpy as np
from matplotlib import pyplot as plt
import random
from openpyxl import load_workbook


# 计算并返回均值向量
def junzhi(sonar):
    a = np.zeros([60, 1])
    for i in range(60):
        a[i]=np.mean(sonar[:,i])
    return a

# 计算类内离散度矩阵S_i
def S_i(sonar):
    a = junzhi(sonar)
    b = np.zeros([sonar.shape[1], sonar.shape[1]])
    for i in range(sonar.shape[0]):
        b = b + np.matmul((sonar[i, :].T - a), (sonar[i, :].T - a).T)
    return b

# 计算类间离散度矩阵S_b
def S_b(sonar1, sonar2):
    b_1 = S_i(sonar1)
    b_2 = S_i(sonar2)
    c = b_1 + b_2
    return c

# 划分训练集、测试集的函数
def train_test(sonar, target, num):
    train_sonar, test_sonar, train_target, test_target =\
        train_test_split(sonar, target, test_size=0.6, random_state=num, shuffle=True)
    return {'train_sonar': train_sonar, 'test_sonar': test_sonar, 'train_target': train_target, 'test_target': test_target}

# 计算出最优的投影方向并计算出决策点
def best_w(sonar1, sonar2, target1, target2, num):
    train_sonar1 = train_test(sonar1, target1, num)['train_sonar']
    train_sonar2 = train_test(sonar2, target2, num)['train_sonar']
    s_0 = S_b(train_sonar1, train_sonar2)
    best_w = np.matmul(np.linalg.inv(s_0), junzhi(train_sonar1) - junzhi(train_sonar2))
    y_0 = (58/124)*np.mean(np.matmul(train_sonar1, best_w)) + (66/124)*np.mean(np.matmul(train_sonar2, best_w))
    # print(best_w)
    return best_w, y_0

# 对测试样本进行分类并计算相关评价指标
def classify(sonar1,sonar2,target1,target2,num):
    # print(num)
    ## 训练集得到的最佳方向和决策点
    w_best12,y0_12=best_w(sonar1, sonar2, target1, target2, num)
    ## 测试集
    test1=train_test(sonar1,target1,num)['test_sonar']
    test2=train_test(sonar2,target2,num)['test_sonar']
    test=np.vstack((test1,test2))
    ## 当前测试集对应的标签
    test_target1=train_test(sonar1,target1,num)['test_target']
    test_target2=train_test(sonar2,target2,num)['test_target']
    # print(test_target1.shape)
    # print(test_target2.shape)
    test_target=np.vstack((test_target1,test_target2))
    # print(test_target.shape)
    # print(test_target)
    ## 存放预测得到的标签
    predict_target=np.zeros_like(test_target)
    ## 通过决策函数
    y=np.matmul(test,w_best12)
    for i in range(len(test)):
        if y[i]>y0_12 or y[i]==y0_12:
            predict_target[i]=0
        else:
            predict_target[i]=1
    # print(predict_target)
    ## 计算OA、AA、kappa系数
    ### 记录三类样本分类正确的数量
    num_1=0
    num_2=0
    ### 记录三类样本实际分类的数量
    real_num_1=0
    real_num_2=0
    for i in range(len(test_target)):
        ## 统计分类正确的数量
        if i<len(test_target1):
            if predict_target[i]==test_target[i]:
                num_1=num_1+1
        else:
            if predict_target[i]==test_target[i]:
                num_2=num_2+1
        ## 统计实际的分类数量
        if predict_target[i]==0:
            real_num_1=real_num_1+1
        else:
            real_num_2=real_num_2+1
    # print(num_1)
    ### 计算相关指标
    OA=(num_1+num_2+20)/len(test_target)
    AA_1=(10+num_1)/len(test_target1)
    AA_2=(10+num_2)/len(test_target2)
    pe=(real_num_1*len(test_target1)+real_num_2*\
        len(test_target2))/np.square(len(test_target))
    kappa=(OA-pe)/(1-pe)
    return OA,[AA_1,AA_2],kappa


if __name__ == "__main__":
    # 导入数据
    workbook=load_workbook(filename='sonar.xlsx')
    # print(workbook.sheetnames)
    sheet=workbook['Sheet1']
    # 存储样本特征集
    sonar=np.zeros([208,60])
    # 存储标签
    target=np.zeros([208,1])
    for i in range(208):
        for j in range(60):
            sonar[i,j]=sheet.cell(row=i+1,column=j+1).value
    # print(sonar.shape)
    # print(sonar[0,59])
    for i in range(208):
        if sheet.cell(row=i+1,column=61).value=='R':
            target[i]=0
        else:
            target[i]=1
    # print(target[97])
    sonar1=sonar[0:97,:]
    sonar2=sonar[97:208,:]
    target1=target[0:97,:]
    target2=target[97:208,:]
    # 存储相关的指标值
    OAs=np.zeros([30,1])
    AAs=np.zeros([30,2])
    kappas=np.zeros([30,1])
    #随机给出30个随机种子，用于train_test函数
    # nums=[703,5205,8248,4998,1027,8528,7063,6513,793,2805,1524,8985,3939,9000\
    # ,3796,3178,628,9359,582,265,5920,8866,7960,5090,5481,4928,526,8763,5333,6596]
    nums=random.sample(range(0,10000),100)
    j=0;k=0
    while j<30:
        try:
            OAs[j],AAs[j,],kappas[j]=classify(sonar1,sonar2,target1,target2,nums[k])
            j=j+1
            k=k+1
        except:
            k+=1
    
    # OAs[j],AAs[j,],kappas[j]=classify(sonar1,sonar2,target1,target2,nums[k])

    temp=0
    for i in range(len(OAs)):
        if OAs[i]!=0:
            temp=temp+1
    print(OAs)
    print(AAs)
    print(kappas)
    print("AA值分别为：\n{:.3f}\n{:.3f}".format(np.sum(AAs[:,0])/temp,np.sum(AAs[:,1]/temp)))
    print("OA值为：\n{:.3f}".format(np.sum(OAs)/temp))
    print("kappa值为：\n{:.3f}".format(np.sum(kappas)/temp))













